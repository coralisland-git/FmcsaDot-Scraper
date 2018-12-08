from selenium import webdriver

from openpyxl import Workbook 

from openpyxl import load_workbook

import os

import platform

import time

import pdb


def is_exist(item, file_name_list):

    ret_flag = False

    item_name = 'USDOT_'+item+'_All_BASICs_Public'

    for file_name in file_name_list:

        if item_name.lower() in file_name.lower():

            ret_flag = True

            break

    return ret_flag


def start_app():   

    dir_path = os.path.dirname(os.path.realpath(__file__))

    downloads_directory = dir_path + '\\downloads'

    try: 

        os.makedirs(downloads_directory)

    except OSError:

        if not os.path.isdir(downloads_directory):

            raise

    exist_file_list = os.listdir(downloads_directory)

    exist_file_list = [file_name for file_name in exist_file_list if '~$' not in file_name]

    excel_file_list = os.listdir('./')

    input_file = 'data.xlsx'

    for file_name in excel_file_list:

        if file_name.endswith('.xlsx'):

            input_file = file_name

            break

    print('Loading Data...')

    key_list = []

    try:

        workbook = load_workbook('./'+input_file )

        mainbook = workbook[workbook.sheetnames[0]]

        for row in mainbook.rows:

            key = str(row[0].value)

            if not is_exist(key, exist_file_list):

                key_list.append(key)

    except Exception as e:

        print('Can not import data.xlsx file. Please check and try again.')  

    chrome_options = webdriver.ChromeOptions()

    # chrome_options.add_argument("headless")

    chrome_options.add_experimental_option("prefs", {"download.default_directory": downloads_directory})

    system = platform.system()

    if system == 'Windows':

        driver = webdriver.Chrome('./chromedriver.exe', chrome_options=chrome_options)

    else:

        driver = webdriver.Chrome('./chromedriver', chrome_options=chrome_options)

    driver.get('https://ai.fmcsa.dot.gov/SMS/Carrier/1165242/Overview.aspx?FirstView=True') 

    total_count = 0

    limit = 5

    while len(key_list) != 0:

        for key in key_list:

            try:

                driver.find_element_by_name('MCSearch').clear()   

                driver.find_element_by_name('MCSearch').send_keys(key)

                time.sleep(2)

                driver.find_elements_by_class_name('ui-menu-item-wrapper')[0].click()

                time.sleep(2)

                driver.find_element_by_xpath('//aside[@id="basic-toolbox-container"]//li[@class="downloads"]').click()

                time.sleep(2)

                driver.find_element_by_xpath('//aside[@id="Downloads"]//input[@value="Download"]').submit()

                time.sleep(2)

                driver.find_element_by_xpath('//a[@class="modalCloseImg simplemodal-close"]').click()

                time.sleep(2)

                key_list.remove(key)

                total_count += 1

                print('USDOT_'+key+'_All_BASICs_Public')

            except Exception as e:

                pass

        limit -= 1

        if limit == 0:

            break

    print('\n~~~~~~~~~~~~~~~~~~~~~~~~~~~\n')

    print('New : ' + str(total_count)+' files downloaded')

    print('Total : ' + str(total_count + len(exist_file_list))+' files')

    print('Does not exist files: '+','.join(key_list))

def main():

    print('Start Downloading...')

    start_app()

    print('\nFinished Successfully')


if __name__ == "__main__":

    main()